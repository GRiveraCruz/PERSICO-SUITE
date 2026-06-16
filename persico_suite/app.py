"""
Persico Mex — Suite Unificada
===============================
Combina:  Job Register         (orig. puerto 5001)
          Hourly Rate Register  (orig. puerto 5002)
          GERC Quote Register   (orig. puerto 5000)
          Purchase Orders       (nuevo módulo)

Ejecutar:  python app.py
Acceso:    http://<IP-del-servidor>:5000
"""

import io, json, re, datetime, shutil
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, Response
from threading import Lock
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════
#  CONFIG  — rutas locales (no depende del servidor NAS)
# ══════════════════════════════════════════════════════════════════
import os as _os
_BASE = _os.path.dirname(_os.path.abspath(__file__))
_DATA = _os.path.join(_BASE, "data")

JOBS_FOLDER  = _os.path.join(_DATA, "JOBs")
RATES_FOLDER = _os.path.join(_DATA, "HOUR_RATE")
XLSM_PATH    = _os.path.join(_DATA, "QUOTE_REG", "quotes.json")   # migrado a JSON
QUOTE_BASE   = _os.path.join(_DATA, "QUOTE_REG")
PO_FOLDER    = _os.path.join(_DATA, "IPOs")
FX_FOLDER    = _os.path.join(_DATA, "FX")

HOST         = "0.0.0.0"
PORT         = 5000
CURRENT_YEAR = datetime.date.today().year

# Quote Register constants
QUOTE_DATA_ROW = 4
QUOTE_MAX_ROWS = 200
# ══════════════════════════════════════════════════════════════════

app  = Flask(__name__, static_folder="static", static_url_path="/static")
lock = Lock()
JOB_RE = re.compile(r"^\d+-\d+$")

# ══════════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ══════════════════════════════════════════════════════════════════
def to_str(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool): return v
    return str(v).strip() or None

def esc_csv(s):
    return str(s or "").replace(",", "")

# ══════════════════════════════════════════════════════════════════
#  JOB REGISTER HELPERS
# ══════════════════════════════════════════════════════════════════
def validate_subindex(sub):
    try:
        n = int(sub)
    except ValueError:
        return False
    return (n == 0 or n == 1 or (2 <= n <= 50) or
            (51 <= n <= 60) or (61 <= n <= 97) or n == 99)

def subindex_label(sub):
    try:
        n = int(sub)
    except ValueError:
        return "Desconocido"
    if n == 0:           return "Máquina / equipo principal"
    if n == 1:           return "Instalación y puesta en marcha"
    if 2  <= n <= 50:    return f"Cambio de ingeniería ({n:02d})"
    if 51 <= n <= 60:    return f"Refacción pagada por cliente ({n})"
    if 61 <= n <= 97:    return f"Servicio pagado por cliente ({n})"
    if n == 99:          return "Servicio de garantía"
    return "Índice no válido"

def jobs_root(): return Path(JOBS_FOLDER)
def job_folder(job_number): return jobs_root() / job_number
def meta_path(job_number): return job_folder(job_number) / "job_info.json"

def read_meta(job_number):
    mp = meta_path(job_number)
    if mp.exists():
        try:
            with open(mp, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    parts = job_number.split("-")
    sub = parts[1] if len(parts) > 1 else "00"
    return {
        "job_number": job_number,
        "main_index": int(parts[0]) if parts[0].isdigit() else 0,
        "subindex": sub.zfill(2),
        "subindex_label": subindex_label(sub),
        "customer": "", "pm": "", "description": "",
        "product_group": "", "product_subgroup": "",
        "revenue": 0, "estimated_cost": 0,
        "po_number": "", "ship_date": "",
        "approval_fc": "ToApprove", "status": "Open",
        "notes": "", "created_at": "",
    }

def write_meta(job_number, data):
    with open(meta_path(job_number), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

def scan_jobs():
    root = jobs_root()
    if not root.exists(): return []
    result = []
    for item in sorted(root.iterdir()):
        if item.is_dir() and JOB_RE.match(item.name):
            result.append(read_meta(item.name))
    result.sort(key=lambda j: (j.get("main_index", 0), int(j.get("subindex", "0"))))
    return result

def next_main_index():
    root = jobs_root()
    if not root.exists(): return 100
    indices = []
    for item in root.iterdir():
        if item.is_dir() and JOB_RE.match(item.name):
            try: indices.append(int(item.name.split("-")[0]))
            except ValueError: pass
    return max(indices) + 1 if indices else 100

def all_job_numbers():
    root = jobs_root()
    if not root.exists(): return set()
    return {item.name for item in root.iterdir()
            if item.is_dir() and JOB_RE.match(item.name)}

def extract_customer(full_addr):
    if not full_addr: return ""
    s = str(full_addr).strip()
    m = re.match(r'^([A-Z][A-Z &]+)-', s)
    if m: return m.group(1).strip()
    return re.split(r'[,\n]', s)[0].strip()[:60]

# ══════════════════════════════════════════════════════════════════
#  HOURLY RATE HELPERS
# ══════════════════════════════════════════════════════════════════
def rates_root(): return Path(RATES_FOLDER)
def rates_file(year): return rates_root() / f"rates_{year}.json"

def load_rates(year):
    p = rates_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_rates(year, records):
    root = rates_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(rates_file(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def available_years():
    root = rates_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^rates_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def normalize_name(name):
    return re.sub(r"\s+", " ", str(name).strip().upper())

# ══════════════════════════════════════════════════════════════════
#  QUOTE REGISTER HELPERS
# ══════════════════════════════════════════════════════════════════
# (Quote Register migrado a JSON — sin dependencia de .xlsm)

def _int_or_none(v):
    try: return int(v) if v not in (None, "", "0", 0) else None
    except: return None

def _gen_qnum(records):
    seq = len(records) + 1
    return f"Q-{datetime.date.today().year}-{seq:03d}"

def _quotes_path():
    p = Path(QUOTE_BASE)
    p.mkdir(parents=True, exist_ok=True)
    return p / "quotes.json"

def _load_quotes():
    p = _quotes_path()
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_quotes(records):
    with open(_quotes_path(), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def read_quote_records():
    with lock:
        records = _load_quotes()
        for i, r in enumerate(records):
            r["row"] = i   # row = índice lógico, compatible con la API existente
        return records

def write_quote_record(data, target_row=None):
    with lock:
        records = _load_quotes()
        if target_row is None:
            qnum = data.get("qnum") or _gen_qnum(records)
            rec = {
                "qnum":       qnum,
                "customer":   data.get("customer", ""),
                "desc":       data.get("desc", ""),
                "machine":    _int_or_none(data.get("machine")),
                "tool":       _int_or_none(data.get("tool")),
                "machTool":   _int_or_none(data.get("machTool")),
                "robotic":    _int_or_none(data.get("robotic")),
                "service":    _int_or_none(data.get("service")),
                "rfq":        data.get("rfq") or None,
                "received":   data.get("received") or None,
                "done":       bool(data.get("done")),
                "sentMgmt":   data.get("sentMgmt") or None,
                "sentClient": data.get("sentClient") or None,
                "notes":      data.get("notes") or None,
                "awarded":    bool(data.get("awarded")),
                "created_at": datetime.datetime.now().isoformat(),
            }
            records.append(rec)
            idx = len(records) - 1
        else:
            idx  = target_row
            if idx < 0 or idx >= len(records):
                raise ValueError(f"Fila {target_row} fuera de rango")
            rec  = records[idx]
            qnum = data.get("qnum", rec.get("qnum"))
            rec.update({
                "qnum":       qnum,
                "customer":   data.get("customer", rec.get("customer", "")),
                "desc":       data.get("desc", rec.get("desc", "")),
                "machine":    _int_or_none(data.get("machine")),
                "tool":       _int_or_none(data.get("tool")),
                "machTool":   _int_or_none(data.get("machTool")),
                "robotic":    _int_or_none(data.get("robotic")),
                "service":    _int_or_none(data.get("service")),
                "rfq":        data.get("rfq") or None,
                "received":   data.get("received") or None,
                "done":       bool(data.get("done")),
                "sentMgmt":   data.get("sentMgmt") or None,
                "sentClient": data.get("sentClient") or None,
                "notes":      data.get("notes") or None,
                "awarded":    bool(data.get("awarded")),
                "updated_at": datetime.datetime.now().isoformat(),
            })
        _save_quotes(records)
        rec["row"] = idx
        return rec

def delete_quote_record(target_row):
    with lock:
        records = _load_quotes()
        if 0 <= target_row < len(records):
            records.pop(target_row)
            _save_quotes(records)

# ══════════════════════════════════════════════════════════════════
#  ROUTES — GENERAL
# ══════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/api/ping")
def ping():
    jobs_ok   = jobs_root().exists()
    rates_ok  = rates_root().exists()
    quotes_file = _quotes_path()
    xlsm_ok   = quotes_file.exists()
    quote_ok  = Path(QUOTE_BASE).exists()
    po_ok     = Path(PO_FOLDER).exists()
    job_count = 0
    if jobs_ok:
        job_count = sum(1 for f in jobs_root().iterdir()
                        if f.is_dir() and JOB_RE.match(f.name))
    return jsonify({
        "jobs_folder":  JOBS_FOLDER,
        "jobs_ok":      jobs_ok,
        "job_count":    job_count,
        "rates_folder": RATES_FOLDER,
        "rates_ok":     rates_ok,
        "years":        available_years(),
        "current_year": CURRENT_YEAR,
        "xlsm_path":    str(quotes_file),
        "xlsm_ok":      xlsm_ok,
        "quote_base":   QUOTE_BASE,
        "quote_ok":     quote_ok,
        "po_folder":    PO_FOLDER,
        "po_ok":        po_ok,
        "wh_folder":    WH_FOLDER,
        "wh_ok":        Path(WH_FOLDER).exists(),
        "ivp_folder":   IVP_FOLDER,
        "ivp_ok":       Path(IVP_FOLDER).exists(),
        "fx_folder":    FX_FOLDER,
        "fx_ok":        Path(FX_FOLDER).exists(),
    })

# ══════════════════════════════════════════════════════════════════
#  ROUTES — JOB REGISTER  (/api/jobs/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/jobs", methods=["GET"])
def api_get_jobs():
    try: return jsonify(scan_jobs())
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/next-index", methods=["GET"])
def api_next_index():
    try: return jsonify({"next": next_main_index()})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/jobs", methods=["POST"])
def api_create_job():
    try:
        data = request.json
        sub  = str(data.get("subindex", "00")).zfill(2)
        if not validate_subindex(sub):
            return jsonify({"error": f"Subíndice '{sub}' no válido."}), 400
        with lock:
            main = next_main_index()
            job_number = f"{main}-{sub}"
            if job_number in all_job_numbers():
                return jsonify({"error": f"El Job {job_number} ya existe."}), 409
            folder = job_folder(job_number)
            try: folder.mkdir(parents=True, exist_ok=True)
            except Exception as fe:
                return jsonify({"error": f"No se pudo crear carpeta en NAS: {fe}"}), 500
            record = {
                "job_number": job_number, "main_index": main,
                "subindex": sub, "subindex_label": subindex_label(sub),
                "customer": data.get("customer", ""),
                "pm": data.get("pm", ""),
                "description": data.get("description", ""),
                "product_group": data.get("product_group", ""),
                "product_subgroup": data.get("product_subgroup", ""),
                "revenue": data.get("revenue", 0),
                "estimated_cost": data.get("estimated_cost", 0),
                "po_number": data.get("po_number", ""),
                "ship_date": data.get("ship_date", ""),
                "approval_fc": data.get("approval_fc", "ToApprove"),
                "status": data.get("status", "Open"),
                "notes": data.get("notes", ""),
                "created_at": datetime.datetime.now().isoformat(),
            }
            write_meta(job_number, record)
            return jsonify(record), 201
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<job_number>", methods=["PUT"])
def api_update_job(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    try:
        data = request.json
        with lock:
            if not job_folder(job_number).exists():
                return jsonify({"error": "Job no encontrado"}), 404
            meta = read_meta(job_number)
            for k in ["customer","pm","description","product_group","product_subgroup",
                      "revenue","estimated_cost","po_number","ship_date",
                      "approval_fc","status","notes"]:
                if k in data: meta[k] = data[k]
            meta["updated_at"] = datetime.datetime.now().isoformat()
            write_meta(job_number, meta)
            return jsonify(meta)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<job_number>", methods=["DELETE"])
def api_delete_job(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    try:
        mp = meta_path(job_number)
        if mp.exists(): mp.unlink()
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/files/<job_number>", methods=["GET"])
def api_list_job_files(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    folder = job_folder(job_number)
    if not folder.exists(): return jsonify([])
    files = []
    for f in sorted(folder.iterdir()):
        if f.is_file() and f.name != "job_info.json":
            st = f.stat()
            files.append({
                "name": f.name, "size": st.st_size,
                "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return jsonify(files)

@app.route("/api/files/<job_number>", methods=["POST"])
def api_upload_job_file(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    folder = job_folder(job_number)
    try: folder.mkdir(parents=True, exist_ok=True)
    except Exception as e: return jsonify({"error": f"No se pudo acceder a la carpeta: {e}"}), 500
    saved = []
    for f in request.files.getlist("files"):
        dest = folder / f.filename
        f.save(str(dest))
        saved.append({"name": f.filename, "size": dest.stat().st_size})
    return jsonify({"saved": saved})

@app.route("/api/files/<job_number>/<filename>", methods=["DELETE"])
def api_delete_job_file(job_number, filename):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    target = job_folder(job_number) / filename
    if target.exists() and target.is_file() and target.name != "job_info.json":
        target.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404

@app.route("/api/import-jobs-excel", methods=["POST"])
def api_import_jobs_excel():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value: headers[str(cell.value).strip()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a in headers: return headers[a]
            return None

        ci_job  = col("Job Number", "Job Sequnce#", "Job #")
        ci_pm   = col("PM Assig.", "PM", "PM Assigned")
        ci_desc = col("Job Description", "Description")
        ci_cust = col("Customer and Ship To:", "Customer", "Customer/Ship To")
        ci_rev  = col("Revenue Amount:", "Revenue Amount", "Revenue")
        ci_cost = col("Estimated Cost:", "Estimated Cost", "Cost")
        ci_fc   = col("Approval By FC", "Approval FC", "FC")
        ci_pg   = col("Product Group")
        ci_psg  = col("Product SubGroup", "Product Subgroup")
        ci_po   = col("PO Number", "PO #")
        ci_ship = col("Ship Date")
        ci_date = col("Date Created", "Created")
        ci_note = col("Notes")

        if ci_job is None:
            return jsonify({"error": "No se encontró la columna 'Job Number' en el Excel"}), 400

        year_filter = request.form.get("year", "")
        try: year_filter = int(year_filter) if year_filter else None
        except ValueError: year_filter = None

        def cv(row_vals, idx):
            if idx is None or idx >= len(row_vals): return None
            return row_vals[idx]
        def ts(v): return str(v).strip() if v is not None else ""
        def tf(v):
            try: return float(v) if v is not None else 0
            except: return 0
        def td(v):
            if v is None: return ""
            if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
            return str(v)[:10]

        existing = all_job_numbers()
        results  = {"created": [], "skipped": [], "errors": []}

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            job_number = ts(cv(row, ci_job)).strip()
            if not job_number or not JOB_RE.match(job_number): continue
            if year_filter and ci_date is not None:
                raw_date = cv(row, ci_date)
                if raw_date and hasattr(raw_date, "year"):
                    if raw_date.year < year_filter:
                        results["skipped"].append({"job": job_number, "reason": f"Año {raw_date.year} < {year_filter}"})
                        continue
            if job_number in existing:
                results["skipped"].append({"job": job_number, "reason": "Ya existe"})
                continue
            parts = job_number.split("-")
            sub = parts[1].zfill(2) if len(parts) > 1 else "00"
            meta = {
                "job_number": job_number,
                "main_index": int(parts[0]) if parts[0].isdigit() else 0,
                "subindex": sub, "subindex_label": subindex_label(sub),
                "customer": extract_customer(ts(cv(row, ci_cust))),
                "customer_full": ts(cv(row, ci_cust)),
                "pm": ts(cv(row, ci_pm)),
                "description": ts(cv(row, ci_desc)),
                "product_group": ts(cv(row, ci_pg)),
                "product_subgroup": ts(cv(row, ci_psg)),
                "revenue": tf(cv(row, ci_rev)),
                "estimated_cost": tf(cv(row, ci_cost)),
                "po_number": ts(cv(row, ci_po)),
                "ship_date": td(cv(row, ci_ship)),
                "approval_fc": ts(cv(row, ci_fc)) or "ToApprove",
                "status": "Open", "notes": ts(cv(row, ci_note)),
                "created_at": td(cv(row, ci_date)), "imported": True,
            }
            with lock:
                folder = job_folder(job_number)
                try:
                    folder.mkdir(parents=True, exist_ok=True)
                    write_meta(job_number, meta)
                    existing.add(job_number)
                    results["created"].append(job_number)
                except Exception as fe:
                    results["errors"].append({"job": job_number, "error": str(fe)})

        results["summary"] = {
            "created": len(results["created"]),
            "skipped": len(results["skipped"]),
            "errors":  len(results["errors"]),
        }
        return jsonify(results)
    except Exception as e: return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  ROUTES — HOURLY RATE  (/api/rates/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/rates", methods=["GET"])
def api_get_rates():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = load_rates(year)
        return jsonify({"year": year, "records": data, "available_years": available_years()})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates", methods=["POST"])
def api_save_rates():
    try:
        payload = request.json
        year    = int(payload.get("year", CURRENT_YEAR))
        records = payload.get("records", [])
        for r in records:
            if not r.get("employee"):
                return jsonify({"error": "Todos los registros deben tener un nombre de empleado"}), 400
            try: float(r["rate"])
            except (ValueError, TypeError):
                return jsonify({"error": f"Tarifa inválida para {r.get('employee')}"}), 400
        with lock: save_rates(year, records)
        return jsonify({"ok": True, "year": year, "count": len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates/employee", methods=["PUT"])
def api_update_employee():
    try:
        data     = request.json
        year     = int(data.get("year", CURRENT_YEAR))
        employee = str(data.get("employee", "")).strip()
        rate     = float(data.get("rate", 0))
        dept     = str(data.get("department", "")).strip()
        notes    = str(data.get("notes", "")).strip()
        if not employee: return jsonify({"error": "Nombre de empleado requerido"}), 400
        with lock:
            records = load_rates(year)
            norm = normalize_name(employee)
            found = False
            for r in records:
                if normalize_name(r["employee"]) == norm:
                    r["rate"] = rate; r["department"] = dept; r["notes"] = notes
                    r["updated_at"] = datetime.datetime.now().isoformat()
                    found = True; break
            if not found:
                records.append({
                    "employee": employee, "rate": rate,
                    "department": dept, "notes": notes,
                    "created_at": datetime.datetime.now().isoformat(),
                })
            save_rates(year, records)
            return jsonify({"ok": True, "found": found, "records": records})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates/employee", methods=["DELETE"])
def api_delete_employee():
    try:
        data = request.json
        year = int(data.get("year", CURRENT_YEAR))
        norm = normalize_name(str(data.get("employee", "")).strip())
        with lock:
            records = load_rates(year)
            before  = len(records)
            records = [r for r in records if normalize_name(r["employee"]) != norm]
            save_rates(year, records)
        return jsonify({"ok": True, "removed": before - len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates/copy-year", methods=["POST"])
def api_copy_year():
    try:
        data        = request.json
        source_year = int(data.get("source_year"))
        target_year = int(data.get("target_year"))
        if source_year == target_year:
            return jsonify({"error": "El año origen y destino deben ser distintos"}), 400
        with lock:
            src = load_rates(source_year)
            if not src: return jsonify({"error": f"No hay tarifas para {source_year}"}), 404
            if rates_file(target_year).exists():
                return jsonify({"error": f"Ya existe una tabla para {target_year}. Elimínala primero."}), 409
            new_records = [{
                "employee": r["employee"], "rate": r["rate"],
                "department": r.get("department", ""), "notes": r.get("notes", ""),
                "copied_from": source_year, "created_at": datetime.datetime.now().isoformat(),
            } for r in src]
            save_rates(target_year, new_records)
        return jsonify({"ok": True, "count": len(new_records), "target_year": target_year})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/import-rates-excel", methods=["POST"])
def api_import_rates_excel():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "replace")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value: headers[str(cell.value).strip().upper()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
            return None

        ci_emp  = col("EMPLOYEE", "NOMBRE", "NAME", "EMPLEADO")
        ci_rate = col("HOURLY RATE", "RATE", "TARIFA", "HOURLY_RATE", "HR RATE")
        ci_dept = col("DEPARTMENT", "DEPT", "DEPARTAMENTO", "AREA")
        ci_note = col("NOTES", "NOTE", "NOTAS", "NOTA")

        if ci_emp is None or ci_rate is None:
            return jsonify({"error": "No se encontraron columnas EMPLOYEE / HOURLY RATE en el Excel"}), 400

        imported = []; errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            emp = str(row[ci_emp]).strip() if row[ci_emp] is not None else ""
            if not emp or emp.upper() == "NONE": continue
            try: rate = float(row[ci_rate]) if row[ci_rate] is not None else 0
            except (ValueError, TypeError):
                errors.append({"employee": emp, "error": "Tarifa no numérica"}); continue
            dept  = str(row[ci_dept]).strip() if ci_dept is not None and row[ci_dept] else ""
            notes = str(row[ci_note]).strip() if ci_note is not None and row[ci_note] else ""
            imported.append({
                "employee": emp, "rate": rate, "department": dept, "notes": notes,
                "imported": True, "created_at": datetime.datetime.now().isoformat(),
            })

        if not imported: return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = load_rates(year)
                existing_map = {normalize_name(r["employee"]): r for r in existing}
                for rec in imported:
                    key = normalize_name(rec["employee"])
                    if key in existing_map:
                        existing_map[key]["rate"] = rec["rate"]
                        existing_map[key]["department"] = rec["department"] or existing_map[key].get("department","")
                        existing_map[key]["updated_at"] = rec["created_at"]
                    else:
                        existing_map[key] = rec
                final = list(existing_map.values())
            save_rates(year, final)

        return jsonify({
            "ok": True, "year": year, "mode": mode,
            "imported": len(imported), "total": len(final), "errors": errors,
        })
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/export-rates/<int:year>")
def api_export_rates(year):
    records = load_rates(year)
    lines   = ["EMPLOYEE,HOURLY RATE,DEPARTMENT,NOTES"]
    for r in records:
        lines.append(f"{esc_csv(r.get('employee',''))},{r.get('rate',0)},{esc_csv(r.get('department',''))},{esc_csv(r.get('notes',''))}")
    return Response("\n".join(lines), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=hourly_rates_{year}.csv"})

# ══════════════════════════════════════════════════════════════════
#  ROUTES — QUOTE REGISTER  (/api/quotes/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/quotes", methods=["GET"])
def api_get_quotes():
    try: return jsonify(read_quote_records())
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes", methods=["POST"])
def api_create_quote():
    try:
        data   = request.json
        result = write_quote_record(data, target_row=None)
        return jsonify(result), 201
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes/<int:row>", methods=["PUT"])
def api_update_quote(row):
    try:
        data   = request.json
        result = write_quote_record(data, target_row=row)
        return jsonify(result)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes/<int:row>", methods=["DELETE"])
def api_delete_quote(row):
    try:
        delete_quote_record(row)
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes/upload/<qnum>", methods=["POST"])
def api_upload_quote(qnum):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    folder = Path(QUOTE_BASE) / qnum
    try: folder.mkdir(parents=True, exist_ok=True)
    except Exception as e: return jsonify({"error": f"No se pudo acceder a la carpeta: {e}"}), 500
    saved = []
    for f in request.files.getlist("files"):
        dest = folder / f.filename
        f.save(str(dest))
        saved.append({"name": f.filename, "size": dest.stat().st_size})
    return jsonify({"saved": saved})

@app.route("/api/quotes/files/<qnum>", methods=["GET"])
def api_list_quote_files(qnum):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    folder = Path(QUOTE_BASE) / qnum
    if not folder.exists(): return jsonify([])
    files = []
    for f in sorted(folder.iterdir()):
        if f.is_file():
            st = f.stat()
            files.append({
                "name": f.name, "size": st.st_size,
                "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return jsonify(files)

@app.route("/api/quotes/files/<qnum>/<filename>", methods=["DELETE"])
def api_delete_quote_file(qnum, filename):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    target = Path(QUOTE_BASE) / qnum / filename
    if target.exists() and target.is_file():
        target.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404

# ══════════════════════════════════════════════════════════════════
#  PURCHASE ORDERS HELPERS
# ══════════════════════════════════════════════════════════════════
PO_COLS = [
    "clave", "fecha_doc", "entregar_a", "nombre",
    "subtotal", "tipo_cambio", "estatus",
    "descuento_financiero", "pct_descuento", "fecha_recepcion"
]

def po_root(): return Path(PO_FOLDER)
def po_json_file(year): return po_root() / f"po_{year}.json"

def po_available_years():
    root = po_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^po_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def po_load(year):
    p = po_json_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def po_save(year, records):
    root = po_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(po_json_file(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def po_to_str(v):
    if v is None: return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()

def po_to_float(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

# ══════════════════════════════════════════════════════════════════
#  ROUTES — PURCHASE ORDERS  (/api/po/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/po", methods=["GET"])
def api_get_po():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = po_load(year)
        return jsonify({"year": year, "records": data, "available_years": po_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/po/import", methods=["POST"])
def api_import_po_excel():
    """
    Importa Purchase Orders desde el Excel con columnas:
      Clave | Fecha de documento | Entregar a | Nombre |
      Subtotal | Tipo de cambio | Estatus |
      Descuento financiero | Porcentaje de descuento | Fecha de recepción
    mode=replace → reemplaza toda la tabla del año
    mode=merge   → agrega / actualiza sin borrar los que no aparecen
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "replace")

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active

        # Build header map (0-based)
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a.lower() in headers: return headers[a.lower()]
            return None

        ci_clave  = col("clave")
        ci_fdoc   = col("fecha de documento")
        ci_dest   = col("entregar a")
        ci_nombre = col("nombre")
        ci_sub    = col("subtotal")
        ci_tc     = col("tipo de cambio")
        ci_est    = col("estatus")
        ci_desc   = col("descuento financiero")
        ci_pct    = col("porcentaje de descuento financ", "porcentaje de descuento financiero", "porcentaje de descuento")
        ci_frec   = col("fecha de recepción", "fecha de recepcion")

        if ci_clave is None:
            return jsonify({"error": "No se encontró la columna 'Clave' en el Excel"}), 400

        imported = []
        errors   = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            def cv(idx):
                if idx is None or idx >= len(row): return None
                return row[idx]

            clave = cv(ci_clave)
            if clave is None or str(clave).strip() == "": continue

            try:
                clave_int = int(clave)
            except (ValueError, TypeError):
                errors.append({"clave": str(clave), "error": "Clave no numérica"})
                continue

            subtotal = po_to_float(cv(ci_sub))
            tc       = po_to_float(cv(ci_tc)) or 1.0
            desc_fin = po_to_float(cv(ci_desc))
            pct_desc = po_to_float(cv(ci_pct))

            imported.append({
                "clave":               clave_int,
                "fecha_doc":           po_to_str(cv(ci_fdoc)),
                "entregar_a":          po_to_str(cv(ci_dest)),
                "nombre":              po_to_str(cv(ci_nombre)),
                "subtotal":            subtotal,
                "tipo_cambio":         tc,
                "subtotal_mxn":        round(subtotal * tc, 2),
                "estatus":             po_to_str(cv(ci_est)),
                "descuento_financiero":desc_fin,
                "pct_descuento":       pct_desc,
                "fecha_recepcion":     po_to_str(cv(ci_frec)),
            })

        if not imported:
            return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = po_load(year)
                existing_map = {r["clave"]: r for r in existing}
                for rec in imported:
                    existing_map[rec["clave"]] = rec
                final = list(existing_map.values())
            po_save(year, final)

        return jsonify({
            "ok":       True,
            "year":     year,
            "mode":     mode,
            "imported": len(imported),
            "total":    len(final),
            "errors":   errors,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/po/export/<int:year>")
def api_export_po(year):
    """Exporta Purchase Orders del año como CSV."""
    records = po_load(year)
    lines = ["Clave,Fecha Documento,Entregar A,Nombre,Subtotal,Tipo Cambio,Subtotal MXN,Estatus,Desc.Financiero,% Desc,Fecha Recepción"]
    for r in records:
        lines.append(",".join([
            str(r.get("clave", "")),
            r.get("fecha_doc", ""),
            '"' + r.get("entregar_a", "").replace('"', '') + '"',
            '"' + r.get("nombre", "").replace('"', '') + '"',
            str(r.get("subtotal", 0)),
            str(r.get("tipo_cambio", 1)),
            str(r.get("subtotal_mxn", 0)),
            r.get("estatus", ""),
            str(r.get("descuento_financiero", 0)),
            str(r.get("pct_descuento", 0)),
            r.get("fecha_recepcion", ""),
        ]))
    return Response(
        "\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=purchase_orders_{year}.csv"}
    )

@app.route("/api/po/years")
def api_po_years():
    return jsonify({"available_years": po_available_years(), "current_year": CURRENT_YEAR})

# ══════════════════════════════════════════════════════════════════
#  WORK HOURS HELPERS
# ══════════════════════════════════════════════════════════════════
WH_FOLDER  = _os.path.join(_DATA, "WHs")
IVP_FOLDER = _os.path.join(_DATA, "IVPs")

def wh_root():  return Path(WH_FOLDER)
def ivp_root(): return Path(IVP_FOLDER)

def wh_json_file(year):  return wh_root()  / f"wh_{year}.json"
def ivp_json_file(year): return ivp_root() / f"ivp_{year}.json"

def _generic_available_years(root_fn, prefix):
    root = root_fn()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(rf"^{prefix}_(\d{{4}})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def wh_available_years():  return _generic_available_years(wh_root,  "wh")
def ivp_available_years(): return _generic_available_years(ivp_root, "ivp")

def _generic_load(json_file_fn, year):
    p = json_file_fn(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def wh_load(year):  return _generic_load(wh_json_file,  year)
def ivp_load(year): return _generic_load(ivp_json_file, year)

def _generic_save(root_fn, json_file_fn, year, records):
    root = root_fn()
    root.mkdir(parents=True, exist_ok=True)
    with open(json_file_fn(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def wh_save(year, records):  _generic_save(wh_root,  wh_json_file,  year, records)
def ivp_save(year, records): _generic_save(ivp_root, ivp_json_file, year, records)

# ══════════════════════════════════════════════════════════════════
#  ROUTES — WORK HOURS  (/api/wh/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/wh", methods=["GET"])
def api_get_wh():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = wh_load(year)
        return jsonify({"year": year, "records": data, "available_years": wh_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wh/import", methods=["POST"])
def api_import_wh():
    """
    Importa Work Hours desde Excel con columnas:
      cboReports | cboFilterFavorites | ID | Employee | Date Worked |
      Work Code  | Hours | Work Description
    Soporta filtro por fecha_inicio / fecha_fin y mode replace/merge.
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        year       = int(request.form.get("year", CURRENT_YEAR))
        mode       = request.form.get("mode", "replace")
        date_from  = request.form.get("date_from", "")   # YYYY-MM-DD
        date_to    = request.form.get("date_to",   "")   # YYYY-MM-DD

        dt_from = None
        dt_to   = None
        if date_from:
            try: dt_from = datetime.datetime.strptime(date_from[:10], "%Y-%m-%d")
            except: pass
        if date_to:
            try: dt_to = datetime.datetime.strptime(date_to[:10], "%Y-%m-%d")
            except: pass

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        # Use the first sheet (may be named 'Work Hours List')
        ws = wb.active if len(wb.sheetnames) == 1 else wb[wb.sheetnames[0]]
        for sname in wb.sheetnames:
            if "work hours" in sname.lower():
                ws = wb[sname]; break

        # Detect header row — scan first 3 rows for 'Employee'
        header_row = 1
        headers = {}
        for ri in range(1, 4):
            row_vals = [c.value for c in next(ws.iter_rows(min_row=ri, max_row=ri))]
            if any(str(v).strip().lower() == "employee" for v in row_vals if v):
                header_row = ri
                for ci, v in enumerate(row_vals):
                    if v: headers[str(v).strip().lower()] = ci
                break

        def col(*aliases):
            for a in aliases:
                if a.lower() in headers: return headers[a.lower()]
            return None

        ci_id    = col("id")
        ci_emp   = col("employee")
        ci_date  = col("date worked", "date")
        ci_wcode = col("work code")
        ci_hours = col("hours")
        ci_desc  = col("work description", "description")

        if ci_emp is None or ci_date is None or ci_hours is None:
            return jsonify({"error": "No se encontraron columnas requeridas (Employee, Date Worked, Hours)"}), 400

        imported = []
        skipped  = 0
        errors   = []

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row = list(row)
            def cv(idx):
                if idx is None or idx >= len(row): return None
                return row[idx]

            emp = cv(ci_emp)
            if not emp or str(emp).strip() == "": continue

            date_val = cv(ci_date)
            if not isinstance(date_val, (datetime.datetime, datetime.date)):
                skipped += 1; continue

            # Date range filter
            if dt_from and date_val < dt_from: skipped += 1; continue
            if dt_to   and date_val > dt_to:   skipped += 1; continue

            try:
                hours = float(cv(ci_hours)) if cv(ci_hours) is not None else 0
            except (ValueError, TypeError):
                errors.append({"row": str(cv(ci_id)), "error": "Horas no numéricas"}); continue

            imported.append({
                "id":          int(cv(ci_id)) if cv(ci_id) is not None else None,
                "employee":    str(emp).strip(),
                "date_worked": date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10],
                "work_code":   str(cv(ci_wcode) or "").strip(),
                "hours":       hours,
                "description": str(cv(ci_desc) or "").strip(),
            })

        if not imported:
            return jsonify({"error": f"No se encontraron registros válidos (omitidos: {skipped})"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = wh_load(year)
                existing_ids = {r["id"] for r in existing if r.get("id")}
                for rec in imported:
                    if rec.get("id") and rec["id"] in existing_ids:
                        for i, ex in enumerate(existing):
                            if ex.get("id") == rec["id"]:
                                existing[i] = rec; break
                    else:
                        existing.append(rec)
                final = existing
            wh_save(year, final)

        return jsonify({
            "ok":       True,
            "year":     year,
            "mode":     mode,
            "imported": len(imported),
            "skipped":  skipped,
            "total":    len(final),
            "errors":   errors,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wh/export/<int:year>")
def api_export_wh(year):
    records = wh_load(year)
    lines = ["ID,Employee,Date Worked,Work Code,Hours,Description"]
    for r in records:
        lines.append(",".join([
            str(r.get("id", "")),
            '"' + r.get("employee", "").replace('"', '') + '"',
            r.get("date_worked", ""),
            '"' + r.get("work_code", "").replace('"', '') + '"',
            str(r.get("hours", 0)),
            '"' + r.get("description", "").replace('"', '') + '"',
        ]))
    return Response(
        "\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=work_hours_{year}.csv"}
    )

# ══════════════════════════════════════════════════════════════════
#  ROUTES — INVOICED POs  (/api/ivp/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/ivp", methods=["GET"])
def api_get_ivp():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = ivp_load(year)
        return jsonify({"year": year, "records": data, "available_years": ivp_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ivp/import", methods=["POST"])
def api_import_ivp():
    """
    Importa Invoiced POs desde Excel con columnas:
      Clave | Entregar a | Nombre | Subtotal | Estatus |
      Fecha de recepción | Fecha de pago | Documento anterior
    Detecta USD por '(dolares)' en nombre del proveedor.
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "replace")

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active

        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers[str(cell.value).strip().lower().rstrip()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                k = a.lower().rstrip()
                if k in headers: return headers[k]
                # partial match
                for hk in headers:
                    if k in hk or hk in k: return headers[hk]
            return None

        ci_clave = col("clave")
        ci_dest  = col("entregar a")
        ci_nomb  = col("nombre")
        ci_sub   = col("subtotal")
        ci_est   = col("estatus")
        ci_frec  = col("fecha de recepción", "fecha de recepcion")
        ci_fpag  = col("fecha de pago")
        ci_doc   = col("documento anterior")

        if ci_clave is None:
            return jsonify({"error": "No se encontró la columna 'Clave' en el Excel"}), 400

        imported = []
        errors   = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            def cv(idx):
                if idx is None or idx >= len(row): return None
                return row[idx]

            clave = cv(ci_clave)
            if clave is None or str(clave).strip() in ("", "None"): continue

            try:
                clave_int = int(clave)
            except (ValueError, TypeError):
                errors.append({"clave": str(clave), "error": "Clave no numérica"}); continue

            nombre = str(cv(ci_nomb) or "").strip()
            is_usd = "(dolares)" in nombre.lower()

            try:
                subtotal = float(cv(ci_sub)) if cv(ci_sub) is not None else 0.0
            except (ValueError, TypeError):
                errors.append({"clave": clave_int, "error": "Subtotal no numérico"}); continue

            estatus = str(cv(ci_est) or "").strip()
            # Skip obviously bad estatus rows (date leaked into column)
            if estatus and re.match(r"\d{4}-\d{2}-\d{2}", estatus):
                estatus = ""

            def to_date(v):
                if v is None: return ""
                if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
                s = str(v)[:10]
                return s if re.match(r"\d{4}-\d{2}-\d{2}", s) else ""

            doc_ant = cv(ci_doc)
            try:
                doc_ant = int(doc_ant) if doc_ant is not None else None
            except (ValueError, TypeError):
                doc_ant = None

            imported.append({
                "clave":            clave_int,
                "entregar_a":       str(cv(ci_dest) or "").strip(),
                "nombre":           nombre,
                "subtotal":         subtotal,
                "moneda":           "USD" if is_usd else "MXN",
                "estatus":          estatus,
                "fecha_recepcion":  to_date(cv(ci_frec)),
                "fecha_pago":       to_date(cv(ci_fpag)),
                "doc_anterior":     doc_ant,
            })

        if not imported:
            return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = ivp_load(year)
                existing_map = {r["clave"]: r for r in existing}
                for rec in imported:
                    existing_map[rec["clave"]] = rec
                final = list(existing_map.values())
            ivp_save(year, final)

        return jsonify({
            "ok":       True,
            "year":     year,
            "mode":     mode,
            "imported": len(imported),
            "total":    len(final),
            "errors":   errors,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ivp/export/<int:year>")
def api_export_ivp(year):
    records = ivp_load(year)
    lines = ["Clave,Entregar A,Nombre,Subtotal,Moneda,Estatus,Fecha Recepcion,Fecha Pago,Doc Anterior"]
    for r in records:
        lines.append(",".join([
            str(r.get("clave", "")),
            '"' + r.get("entregar_a", "").replace('"', '') + '"',
            '"' + r.get("nombre", "").replace('"', '') + '"',
            str(r.get("subtotal", 0)),
            r.get("moneda", "MXN"),
            r.get("estatus", ""),
            r.get("fecha_recepcion", ""),
            r.get("fecha_pago", ""),
            str(r.get("doc_anterior", "") or ""),
        ]))
    return Response(
        "\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=invoiced_pos_{year}.csv"}
    )

# (run block moved to end)

# ══════════════════════════════════════════════════════════════════
#  JOB REPORT  — /api/report/*
# ══════════════════════════════════════════════════════════════════

def _build_report_data(job_number, rate_year, wh_year, po_year):
    """Core logic: compile all report data for a Job."""
    job_meta = read_meta(job_number) if job_folder(job_number).exists() else {}

    rates_raw = load_rates(rate_year)
    rate_map  = {normalize_name(r["employee"]): float(r["rate"])
                 for r in rates_raw if r.get("employee")}

    wh_raw   = wh_load(wh_year)
    job_main = "-".join(job_number.split("-")[:2]) if "-" in job_number else job_number
    wh_f     = [r for r in wh_raw
                if job_main.upper() in (r.get("work_code") or "").upper()]

    emp_agg = {}
    for r in wh_f:
        emp = str(r.get("employee", "")).strip()
        hrs = float(r.get("hours", 0))
        if emp not in emp_agg:
            emp_agg[emp] = {"employee": emp, "hours": 0.0, "rate": 0.0, "amount": 0.0}
        emp_agg[emp]["hours"] += hrs
        rate = rate_map.get(normalize_name(emp), 0.0)
        emp_agg[emp]["rate"]   = rate
        emp_agg[emp]["amount"] = round(emp_agg[emp]["hours"] * rate, 2)

    workers   = sorted(emp_agg.values(), key=lambda x: x["hours"], reverse=True)
    accum_hrs = round(sum(w["hours"]  for w in workers), 2)
    amount_wh = round(sum(w["amount"] for w in workers), 2)

    po_raw = po_load(po_year)
    po_f   = [r for r in po_raw
              if job_main.upper() in (r.get("entregar_a") or "").upper()]

    fx_all = fx_load_all()
    po_items = [{"clave":       r.get("clave"),
                 "nombre":      r.get("nombre", ""),
                 "subtotal":    float(r.get("subtotal", 0)),
                 "moneda":      r.get("moneda", "MXN"),
                 "subtotal_usd": _po_usd(r, fx_all),
                 "fx_rate":     fx_rate_for_date(r.get("fecha_recepcion",""), fx_all) or float(r.get("tipo_cambio",0)) or None,
                 "estatus":     r.get("estatus", ""),
                 "fecha_recepcion": r.get("fecha_recepcion", "")}
                for r in po_f]

    purch_tot = round(sum(p["subtotal_usd"] for p in po_items), 2)
    revenue   = float(job_meta.get("revenue", 0))
    cost      = round(amount_wh + purch_tot, 2)
    gm        = round(revenue - cost, 2)
    gm_pct    = round((gm / revenue * 100), 1) if revenue else 0.0

    return {
        "job_number":       job_number,
        "customer":         job_meta.get("customer", ""),
        "description":      job_meta.get("description", ""),
        "pm":               job_meta.get("pm", ""),
        "revenue":          revenue,
        "po_number":        job_meta.get("po_number", ""),
        "ship_date":        job_meta.get("ship_date", ""),
        "status":           job_meta.get("status", ""),
        "product_group":    job_meta.get("product_group", ""),
        "accum_hours":      accum_hrs,
        "amount_wh":        amount_wh,
        "workers":          workers,
        "purchasing_total": purch_tot,
        "po_items":         po_items,
        "cost":             cost,
        "gross_margin":     gm,
        "gm_pct":           gm_pct,
        "rate_year":        rate_year,
        "wh_year":          wh_year,
        "po_year":          po_year,
        "wh_matches":       len(wh_f),
        "po_matches":       len(po_f),
    }


@app.route("/api/report/data")
def api_report_data():
    try:
        job_number = request.args.get("job", "").strip()
        rate_year  = int(request.args.get("rate_year", CURRENT_YEAR))
        wh_year    = int(request.args.get("wh_year",   CURRENT_YEAR))
        po_year    = int(request.args.get("po_year",   CURRENT_YEAR))
        if not job_number:
            return jsonify({"error": "job_number requerido"}), 400
        return jsonify(_build_report_data(job_number, rate_year, wh_year, po_year))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/report/export-excel")
def api_report_export_excel():
    """Exporta el reporte como .xlsx siguiendo la estructura del template."""
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    job_number = request.args.get("job", "").strip()
    rate_year  = int(request.args.get("rate_year", CURRENT_YEAR))
    wh_year    = int(request.args.get("wh_year",   CURRENT_YEAR))
    po_year    = int(request.args.get("po_year",   CURRENT_YEAR))

    if not job_number:
        return jsonify({"error": "job_number requerido"}), 400

    d = _build_report_data(job_number, rate_year, wh_year, po_year)

    # Colour palette
    RED_H   = "C8102E"
    DARK    = "1F1F1F"
    DGRAY   = "2D2D2D"
    MGRAY   = "3D3D3D"
    LGRAY   = "F0F0F0"
    XLGRAY  = "FAFAFA"
    GOLD    = "F0A500"
    WHITE   = "FFFFFF"
    GREEN_C = "1E8449"
    RED_NEG = "C0392B"
    BLUE_H  = "1F618D"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _font(sz=9, bold=False, color=DARK, italic=False):
        return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)
    def _side():
        return Side(style="thin", color="AAAAAA")
    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)
    def _lft(indent=1):
        return Alignment(horizontal="left",   vertical="center", indent=indent, wrap_text=False)
    def _rgt():
        return Alignment(horizontal="right",  vertical="center")
    def _ctr():
        return Alignment(horizontal="center", vertical="center")
    MONEY = '#,##0.00'
    HRS   = '#,##0.0'
    PCT   = '0.0"%"'

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report {job_number}"

    # Column widths
    widths = {"A":26,"B":18,"C":28,"D":10,"E":14,"F":16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: Title ─────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"  JOB COST REPORT  ·  {job_number}"
    c.font      = _font(16, True, WHITE)
    c.fill      = _fill(RED_H)
    c.alignment = _lft(2)

    # ── Row 2: Sub-header ─────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    ws.merge_cells("A2:C2")
    ws["A2"].value = f"Generated: {datetime.date.today()}  |  Rate year: {rate_year}  |  WH year: {wh_year}  |  PO year: {po_year}"
    ws["A2"].font  = _font(8, False, "888888", True)
    ws["A2"].alignment = _lft()

    # ── Summary block (rows 3-14) ──────────────────────────────────
    def label_row(row, label, val, fmt=None, bg_lbl=DGRAY, bg_val=MGRAY,
                  fc_lbl=WHITE, fc_val=WHITE, bold_val=False, height=18):
        ws.row_dimensions[row].height = height
        cl = ws.cell(row, 1)
        cl.value = label; cl.font = _font(9, True, fc_lbl)
        cl.fill = _fill(bg_lbl); cl.alignment = _lft(); cl.border = _border()
        cv = ws.cell(row, 2)
        cv.value = val; cv.font = _font(10, bold_val, fc_val)
        cv.fill = _fill(bg_val); cv.alignment = _lft()
        cv.border = _border()
        if fmt: cv.number_format = fmt

    label_row(3,  "JOB NUMBER",            job_number,              bg_val=MGRAY, fc_val=GOLD, bold_val=True)
    label_row(4,  "CUSTOMER",              d["customer"] or "—",    bg_val=MGRAY)
    label_row(5,  "PM",                    d["pm"] or "—",          bg_val=MGRAY)
    label_row(6,  "DESCRIPTION",           d["description"] or "—", bg_val=MGRAY)
    label_row(7,  "STATUS",                d["status"] or "—",      bg_val=MGRAY)

    # Spacer
    ws.row_dimensions[8].height = 6
    ws.merge_cells("A8:F8"); ws["A8"].fill = _fill(DARK)

    label_row(9,  "REVENUE",              d["revenue"],       MONEY, bg_val=BLUE_H, fc_val=WHITE, bold_val=True)
    label_row(10, "ACUMULATED WORK HOURS",d["accum_hours"],   HRS,   bg_val=MGRAY,  fc_val=WHITE)
    label_row(11, "AMOUNT WORK HOURS",    d["amount_wh"],     MONEY, bg_val=MGRAY,  fc_val=WHITE, bold_val=True)
    label_row(12, "PURCHASINGS TOTAL",    d["purchasing_total"], MONEY, bg_val=MGRAY, fc_val=WHITE, bold_val=True)

    ws.row_dimensions[13].height = 6
    ws.merge_cells("A13:F13"); ws["A13"].fill = _fill(DARK)

    # COST
    ws.row_dimensions[14].height = 20
    cl = ws["A14"]; cl.value = "COST"
    cl.font = _font(11, True, WHITE); cl.fill = _fill(RED_H)
    cl.alignment = _lft(); cl.border = _border()
    cv = ws["B14"]; cv.value = d["cost"]
    cv.font = _font(12, True, WHITE); cv.fill = _fill(RED_H)
    cv.alignment = _rgt(); cv.number_format = MONEY; cv.border = _border()

    # GROSS MARGIN
    ws.row_dimensions[15].height = 22
    gm_bg = GREEN_C if d["gross_margin"] >= 0 else RED_NEG
    cl = ws["A15"]; cl.value = "GROSS MARGIN"
    cl.font = _font(12, True, WHITE); cl.fill = _fill(gm_bg)
    cl.alignment = _lft(); cl.border = _border()
    cv = ws["B15"]; cv.value = d["gross_margin"]
    cv.font = _font(13, True, WHITE); cv.fill = _fill(gm_bg)
    cv.alignment = _rgt(); cv.number_format = MONEY; cv.border = _border()

    # GM%
    ws.row_dimensions[16].height = 16
    ws.merge_cells("A16:B16")
    c16 = ws["A16"]
    c16.value = f"Gross Margin %:  {d['gm_pct']:.1f}%"
    c16.font  = _font(10, True, WHITE)
    c16.fill  = _fill(gm_bg); c16.alignment = _ctr(); c16.border = _border()

    # ── Detail tables (right side of summary, rows 3-16) ──────────
    # PO mini-list header (cols D-F, row 3)
    ws.row_dimensions[3].height = max(ws.row_dimensions[3].height, 18)
    for col, txt in [(4,"CLAVE PO"),(5,"MXN"),(6,"PROVEEDOR")]:
        c = ws.cell(3, col)
        c.value = txt; c.font = _font(8, True, WHITE)
        c.fill = _fill(RED_H); c.alignment = _ctr(); c.border = _border()

    for i, po in enumerate(d["po_items"][:12]):
        r = 4 + i
        ws.row_dimensions[r].height = 15
        bg = LGRAY if i % 2 == 0 else XLGRAY
        ws.cell(r,4).value = str(po["clave"]); ws.cell(r,4).font = _font(8,False,"333333")
        ws.cell(r,4).fill = _fill(bg); ws.cell(r,4).alignment = _ctr(); ws.cell(r,4).border = _border()
        ws.cell(r,5).value = po["subtotal_usd"]; ws.cell(r,5).font = _font(8,False,"333333")
        ws.cell(r,5).fill = _fill(bg); ws.cell(r,5).alignment = _rgt()
        ws.cell(r,5).number_format = MONEY; ws.cell(r,5).border = _border()
        ws.cell(r,6).value = po["nombre"][:30]; ws.cell(r,6).font = _font(7,False,"666666")
        ws.cell(r,6).fill = _fill(bg); ws.cell(r,6).alignment = _lft(); ws.cell(r,6).border = _border()

    # ── Spacer ────────────────────────────────────────────────────
    sr = 17
    ws.row_dimensions[sr].height = 8
    ws.merge_cells(f"A{sr}:F{sr}")
    ws[f"A{sr}"].fill = _fill(DARK)

    # ── Detail tables header row ──────────────────────────────────
    dh = sr + 1
    ws.row_dimensions[dh].height = 22
    for col, txt in [(1,"PO NUMBER"),(2,"VALUE (USD)"),(3,"PROVEEDOR"),
                     (4,"WORKER"),(5,"HOURS"),(6,"VALUE (USD)")]:
        c = ws.cell(dh, col)
        c.value = txt; c.font = _font(9, True, WHITE)
        c.fill = _fill(DGRAY); c.alignment = _ctr(); c.border = _border()

    # ── PO detail rows ────────────────────────────────────────────
    po_start = dh + 1
    for i, po in enumerate(d["po_items"]):
        r = po_start + i
        ws.row_dimensions[r].height = 15
        bg = LGRAY if i % 2 == 0 else XLGRAY
        ws.cell(r,1).value = str(po["clave"])
        ws.cell(r,1).font = _font(9,False,"222222"); ws.cell(r,1).fill = _fill(bg)
        ws.cell(r,1).alignment = _ctr(); ws.cell(r,1).border = _border()
        ws.cell(r,2).value = po["subtotal_usd"]
        ws.cell(r,2).font = _font(9,False,"222222"); ws.cell(r,2).fill = _fill(bg)
        ws.cell(r,2).alignment = _rgt(); ws.cell(r,2).number_format = MONEY
        ws.cell(r,2).border = _border()
        ws.cell(r,3).value = po["nombre"][:35]
        ws.cell(r,3).font = _font(8,False,"555555"); ws.cell(r,3).fill = _fill(bg)
        ws.cell(r,3).alignment = _lft(); ws.cell(r,3).border = _border()

    # ── Worker detail rows ────────────────────────────────────────
    wk_start = dh + 1
    for i, w in enumerate(d["workers"]):
        r = wk_start + i
        if r < po_start + len(d["po_items"]):
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height, 15)
        else:
            ws.row_dimensions[r].height = 15
        bg = LGRAY if i % 2 == 0 else XLGRAY
        ws.cell(r,4).value = w["employee"]
        ws.cell(r,4).font = _font(9,False,"222222"); ws.cell(r,4).fill = _fill(bg)
        ws.cell(r,4).alignment = _lft(); ws.cell(r,4).border = _border()
        ws.cell(r,5).value = w["hours"]
        ws.cell(r,5).font = _font(9,False,"222222"); ws.cell(r,5).fill = _fill(bg)
        ws.cell(r,5).alignment = _rgt(); ws.cell(r,5).number_format = HRS
        ws.cell(r,5).border = _border()
        ws.cell(r,6).value = w["amount"]
        ws.cell(r,6).font = _font(9,False,"222222"); ws.cell(r,6).fill = _fill(bg)
        ws.cell(r,6).alignment = _rgt(); ws.cell(r,6).number_format = MONEY
        ws.cell(r,6).border = _border()

    # ── Totals footer ─────────────────────────────────────────────
    foot = max(po_start+len(d["po_items"]), wk_start+len(d["workers"])) + 1
    ws.row_dimensions[foot].height = 20
    for col, val, fmt in [
        (1,"TOTAL",None),(2,d["purchasing_total"],MONEY),(3,"",None),
        (4,"TOTAL",None),(5,d["accum_hours"],HRS),(6,d["amount_wh"],MONEY)]:
        c = ws.cell(foot, col)
        c.value = val; c.font = _font(10, True, WHITE)
        c.fill = _fill(DARK); c.alignment = _rgt() if fmt else _ctr()
        c.border = _border()
        if fmt: c.number_format = fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Report_{job_number.replace('-','_')}_{datetime.date.today()}.xlsx"
    resp  = make_response(buf.read())
    resp.headers["Content-Type"]        = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp

# ══════════════════════════════════════════════════════════════════
#  FX (TIPO DE CAMBIO) — /api/fx/*
# ══════════════════════════════════════════════════════════════════

def fx_root():           return Path(FX_FOLDER)
def fx_json_file(year):  return fx_root() / f"fx_{year}.json"

def fx_available_years():
    root = fx_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^fx_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def fx_load(year) -> dict:
    """Returns {YYYY-MM-DD: rate_float}"""
    p = fx_json_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def fx_save(year, data: dict):
    root = fx_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(fx_json_file(year), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def fx_load_all() -> dict:
    """Merge all years into one lookup dict {YYYY-MM-DD: rate}"""
    combined = {}
    for year in fx_available_years():
        combined.update(fx_load(year))
    return combined

def fx_rate_for_date(date_str: str, fx_all: dict) -> float:
    """
    Returns the MXN/USD rate for a given date string (YYYY-MM-DD).
    Falls back up to 7 days earlier for weekends/holidays.
    Returns None if not found.
    """
    if not date_str or not fx_all:
        return None
    try:
        d = datetime.datetime.strptime(date_str[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    for offset in range(8):
        key = (d - datetime.timedelta(days=offset)).strftime("%Y-%m-%d")
        if key in fx_all:
            return fx_all[key]
    return None


@app.route("/api/fx", methods=["GET"])
def api_get_fx():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = fx_load(year)
        # Return as sorted list for frontend table
        records = [{"date": k, "rate": v} for k, v in sorted(data.items())]
        return jsonify({
            "year":            year,
            "records":         records,
            "available_years": fx_available_years(),
            "count":           len(records),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fx/import", methods=["POST"])
def api_import_fx():
    """
    Importa el archivo tipoCambio.xls del Banco de México.
    Formato: filas de datos a partir de fila 9 (idx 8).
    Col 0 = Fecha (dd/mm/yyyy string)
    Col 3 = Tipo de cambio 'Para solventar obligaciones'
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        mode = request.form.get("mode", "merge")   # merge default — accumulate years
        raw  = f.read()

        # Support both .xls (legacy binary) and .xlsx
        fname_lower = (f.filename or "").lower()

        def parse_rows_from_raw(raw_bytes, is_xls):
            """
            Returns an iterable of rows starting from data row 9 (idx 8).
            For .xls we try xlrd, then fall back to converting via LibreOffice,
            then fall back to a minimal built-in compound-doc reader.
            """
            if is_xls:
                # ── Try xlrd first (if installed) ──────────────────
                try:
                    import xlrd
                    wb2 = xlrd.open_workbook(file_contents=raw_bytes)
                    ws2 = wb2.sheet_by_index(0)
                    return [[ws2.cell_value(r, c) for c in range(ws2.ncols)]
                            for r in range(8, ws2.nrows)]
                except ImportError:
                    pass

                # ── Try pandas with openpyxl-xlrd engine ───────────
                try:
                    import pandas as pd
                    df = pd.read_excel(io.BytesIO(raw_bytes), header=None,
                                       engine='xlrd', skiprows=8)
                    return df.values.tolist()
                except Exception:
                    pass

                # ── Convert .xls → .xlsx via LibreOffice ────────────
                import tempfile, subprocess, os
                with tempfile.TemporaryDirectory() as tmpdir:
                    src = os.path.join(tmpdir, "tc.xls")
                    with open(src, "wb") as fh:
                        fh.write(raw_bytes)
                    result = subprocess.run(
                        ["libreoffice", "--headless", "--convert-to", "xlsx",
                         "--outdir", tmpdir, src],
                        capture_output=True, timeout=30
                    )
                    out_path = src.replace(".xls", ".xlsx")
                    if result.returncode != 0 or not os.path.exists(out_path):
                        raise RuntimeError(
                            "No se pudo convertir el archivo. "
                            "Instala xlrd: pip install xlrd==1.2.0 --break-system-packages"
                        )
                    with open(out_path, "rb") as fh:
                        xlsx_bytes = fh.read()

                wb3 = openpyxl.load_workbook(io.BytesIO(xlsx_bytes),
                                              read_only=True, data_only=True)
                ws3 = wb3.active
                return list(ws3.iter_rows(min_row=9, values_only=True))

            else:
                wb4 = openpyxl.load_workbook(io.BytesIO(raw_bytes),
                                              read_only=True, data_only=True)
                ws4 = wb4.active
                return list(ws4.iter_rows(min_row=9, values_only=True))

        is_xls = fname_lower.endswith(".xls") and not fname_lower.endswith(".xlsx")
        try:
            all_rows = parse_rows_from_raw(raw, is_xls)
        except RuntimeError as re_err:
            return jsonify({"error": str(re_err)}), 400

        rows_iter = all_rows

        # Parse rows
        by_year   = {}    # year → {YYYY-MM-DD: rate}
        imported  = 0
        skipped   = 0

        for row in rows_iter:
            if not row or not row[0]:
                continue
            date_raw = str(row[0]).strip()
            rate_raw = row[3] if len(row) > 3 else None

            # Parse date dd/mm/yyyy
            try:
                d = datetime.datetime.strptime(date_raw, "%d/%m/%Y")
            except ValueError:
                skipped += 1
                continue

            # Parse rate — skip N/E
            try:
                rate = float(rate_raw)
                if rate <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                skipped += 1
                continue

            year    = d.year
            iso_key = d.strftime("%Y-%m-%d")
            by_year.setdefault(year, {})[iso_key] = round(rate, 6)
            imported += 1

        if not by_year:
            return jsonify({"error": "No se encontraron registros válidos"}), 400

        total_saved = 0
        with lock:
            for year, new_data in by_year.items():
                if mode == "replace":
                    final = new_data
                else:   # merge
                    existing = fx_load(year)
                    existing.update(new_data)
                    final = existing
                fx_save(year, final)
                total_saved += len(final)

        years_touched = sorted(by_year.keys())
        return jsonify({
            "ok":       True,
            "imported": imported,
            "skipped":  skipped,
            "years":    years_touched,
            "total_saved": total_saved,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fx/lookup")
def api_fx_lookup():
    """Quick single-date lookup: ?date=YYYY-MM-DD"""
    try:
        date_str = request.args.get("date", "")
        fx_all   = fx_load_all()
        rate     = fx_rate_for_date(date_str, fx_all)
        return jsonify({"date": date_str, "rate": rate, "found": rate is not None})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Helper used by PO and IVP endpoints: convert subtotal to USD ─
def _po_usd(record: dict, fx_all: dict) -> float:
    """Return the subtotal in USD for a PO record."""
    subtotal = float(record.get("subtotal", 0))
    moneda   = record.get("moneda", "MXN")
    if moneda == "USD":
        return subtotal
    # MXN → USD
    date_str = record.get("fecha_recepcion") or record.get("fecha_doc") or ""
    rate     = fx_rate_for_date(date_str, fx_all)
    if rate and rate > 0:
        return round(subtotal / rate, 6)
    # Fallback: use stored tipo_cambio if available (PO module)
    tc = float(record.get("tipo_cambio", 0))
    if tc > 1:
        return round(subtotal / tc, 6)
    return subtotal   # can't convert — return as-is


@app.route("/api/po/usd-view")
def api_po_usd_view():
    """Return PO records with all amounts converted to USD."""
    try:
        year   = int(request.args.get("year", CURRENT_YEAR))
        po_raw = po_load(year)
        fx_all = fx_load_all()
        result = []
        for r in po_raw:
            rec = dict(r)
            rec["subtotal_usd"] = _po_usd(r, fx_all)
            # Determine which rate was used
            if r.get("moneda") == "USD":
                rec["fx_rate_used"] = 1.0
            else:
                date_str = r.get("fecha_recepcion") or r.get("fecha_doc") or ""
                rec["fx_rate_used"] = fx_rate_for_date(date_str, fx_all) or float(r.get("tipo_cambio", 0)) or None
            result.append(rec)
        return jsonify({"year": year, "records": result, "available_years": po_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("  Persico Mex — Suite Unificada")
    print(f"  Job Register    : {JOBS_FOLDER}")
    print(f"  Hourly Rates    : {RATES_FOLDER}")
    print(f"  Quote Register  : {QUOTE_BASE}/quotes.json")
    print(f"  Purchase Orders : {PO_FOLDER}")
    print(f"  Work Hours      : {WH_FOLDER}")
    print(f"  Invoiced POs    : {IVP_FOLDER}")
    print(f"  FX / Tipo cambio: {FX_FOLDER}")
    print(f"  URL             : http://localhost:{PORT}")
    print("=" * 60)
    app.run(host=HOST, port=PORT, debug=False)
